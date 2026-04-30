"""Препроцессинг входного xlsx: парсинг диалогов, эвристическая разметка исхода,
сбор статистики длины (для адаптивного cap по ходам), сэмплинг подмножества
для дальнейшего LLM-анализа.

Формат входного xlsx (наследуется от текущего корпуса СберБизнес/эквайринг):
    Id_диалога | Порядок_реплики | Реплика_Спикер | Реплика_Текст

Спикеры: 'Клиент' | 'Сотрудник'.
"""
from __future__ import annotations

import random
import re
import statistics
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

import openpyxl


SPEAKER_CLIENT = "Клиент"
SPEAKER_AGENT = "Сотрудник"


@dataclass
class Turn:
    order: int
    speaker: str
    text: str


@dataclass
class Dialogue:
    id: str
    turns: list[Turn] = field(default_factory=list)
    outcome: str = "unknown"  # agree | transfer | refuse | off_target | unknown

    @property
    def length(self) -> int:
        return len(self.turns)

    def as_text(self, max_turns: int | None = None) -> str:
        ts = self.turns if max_turns is None else self.turns[:max_turns]
        return "\n".join(f"{t.speaker}: {t.text}" for t in ts)

    def tail_text(self, n: int) -> str:
        return "\n".join(f"{t.speaker}: {t.text}" for t in self.turns[-n:])

    def head_tail_text(self, head: int, tail: int) -> str:
        """Первые `head` + последние `tail` реплик. Короткие диалоги — целиком."""
        if len(self.turns) <= head + tail:
            return self.as_text()
        head_part = "\n".join(f"{t.speaker}: {t.text}" for t in self.turns[:head])
        tail_part = "\n".join(f"{t.speaker}: {t.text}" for t in self.turns[-tail:])
        skipped = len(self.turns) - head - tail
        return f"{head_part}\n... [пропущено {skipped} реплик] ...\n{tail_part}"


@dataclass
class CorpusStats:
    dialogue_count: int
    median_turns: int
    p90_turns: int
    avg_chars_per_turn: float
    outcome_counts: dict[str, int]


# ---------------------------------------------------------------------------
# Парсинг
# ---------------------------------------------------------------------------
def parse_xlsx(path: str | Path) -> list[Dialogue]:
    """Читаем xlsx → список диалогов, отсортированных по `Порядок_реплики`."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return []
    cols = {str(h).strip().lower(): i for i, h in enumerate(header) if h is not None}

    def col(name: str) -> int:
        for k, v in cols.items():
            if name in k:
                return v
        raise ValueError(f"В xlsx не найден столбец, похожий на «{name}»: {list(cols)}")

    i_id = col("id_диалога")
    i_ord = col("порядок")
    i_sp = col("спикер")
    i_tx = col("текст")

    bucket: dict[str, list[Turn]] = defaultdict(list)
    for row in rows:
        if not row or row[i_id] is None:
            continue
        text = row[i_tx]
        if text is None:
            continue
        try:
            order = int(row[i_ord]) if row[i_ord] is not None else len(bucket[str(row[i_id])]) + 1
        except (TypeError, ValueError):
            continue
        bucket[str(row[i_id])].append(
            Turn(order=order, speaker=str(row[i_sp] or "?").strip(), text=str(text).strip())
        )

    dialogues = []
    for did, turns in bucket.items():
        turns.sort(key=lambda t: t.order)
        if turns:
            dialogues.append(Dialogue(id=did, turns=turns))
    return dialogues


# ---------------------------------------------------------------------------
# Эвристическая разметка исхода
# ---------------------------------------------------------------------------
_PATTERNS: list[tuple[str, re.Pattern[str]]] = [
    (
        "off_target",
        re.compile(
            r"не\s+туда|ошиблись\s+номером|вы\s+кому|никаких\s+заявок\s+не\s+оставлял",
            re.IGNORECASE,
        ),
    ),
    (
        "agree",
        re.compile(
            r"\bда[, ]\s*давайте\b|\bсоглас(ен|на)\b|\bоформляйте\b|"
            r"\bдоговорились\b|\bхорошо[, ]\s*записывайте\b|\bподходит\b",
            re.IGNORECASE,
        ),
    ),
    (
        "refuse",
        re.compile(
            r"не\s+интересно|нет[, ]\s*спасибо|не\s+надо|не\s+нужно|"
            r"откажусь|не\s+подходит|не\s+будем",
            re.IGNORECASE,
        ),
    ),
    (
        "transfer",
        re.compile(
            r"перезвоните\s+(позже|потом)|подума(ю|ем)|не\s+сейчас|"
            r"\bпотом\b|обсудим\s+позже|давайте\s+попозже",
            re.IGNORECASE,
        ),
    ),
]


def heuristic_outcome(dialogue: Dialogue, tail_turns: int = 5) -> str:
    tail = dialogue.tail_text(tail_turns)
    for label, pat in _PATTERNS:
        if pat.search(tail):
            return label
    return "unknown"


def mark_outcomes(dialogues: Iterable[Dialogue]) -> None:
    """Проставляем outcome in-place на основе хвоста диалога."""
    for d in dialogues:
        d.outcome = heuristic_outcome(d)


# ---------------------------------------------------------------------------
# Статистика
# ---------------------------------------------------------------------------
def compute_stats(dialogues: list[Dialogue]) -> CorpusStats:
    if not dialogues:
        return CorpusStats(0, 0, 0, 0.0, {})
    lengths = sorted(d.length for d in dialogues)
    median = lengths[len(lengths) // 2]
    p90 = lengths[max(0, int(len(lengths) * 0.9) - 1)]
    chars = sum(len(t.text) for d in dialogues for t in d.turns)
    turns_total = sum(d.length for d in dialogues)
    avg = chars / turns_total if turns_total else 0.0
    return CorpusStats(
        dialogue_count=len(dialogues),
        median_turns=int(median),
        p90_turns=int(p90),
        avg_chars_per_turn=avg,
        outcome_counts=dict(Counter(d.outcome for d in dialogues)),
    )


# ---------------------------------------------------------------------------
# Сэмплинг
# ---------------------------------------------------------------------------
def sample_balanced(
    dialogues: list[Dialogue],
    n: int,
    seed: int = 42,
) -> list[Dialogue]:
    """Берём `n` диалогов, равномерно распределяя по outcome-классам.

    Если в классе диалогов меньше квоты — берём все имеющиеся и добираем из
    общего пула.
    """
    if n >= len(dialogues):
        return list(dialogues)

    rng = random.Random(seed)
    by_outcome: dict[str, list[Dialogue]] = defaultdict(list)
    for d in dialogues:
        by_outcome[d.outcome].append(d)

    classes = list(by_outcome)
    quota = max(1, n // max(1, len(classes)))
    picked: list[Dialogue] = []
    for cls in classes:
        pool = by_outcome[cls]
        rng.shuffle(pool)
        picked.extend(pool[:quota])

    if len(picked) < n:
        rest = [d for d in dialogues if d not in picked]
        rng.shuffle(rest)
        picked.extend(rest[: n - len(picked)])

    return picked[:n]


def adaptive_turn_cap(stats: CorpusStats, user_cap: int) -> int:
    """Сколько ходов жмём в симуляции: min(пользовательский cap, p90 корпуса).

    Если статистики нет — возвращаем cap пользователя.
    """
    if stats.p90_turns <= 0:
        return user_cap
    return min(user_cap, max(8, stats.p90_turns))
