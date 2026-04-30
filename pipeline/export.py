"""Сборка финального .xlsx-отчёта по прогону."""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

from .agents.censor import CensorVerdict
from .agents.hypothesis_generator import Hypothesis
from .agents.persona_extractor import Persona
from .agents.reporter import HypothesisReport
from .agents.simulator import SimResult


def export_run(
    *,
    out_path: Path,
    hypotheses: list[Hypothesis],
    personas: list[Persona],
    sims: list[SimResult],
    verdicts: list[CensorVerdict],
    reports: list[HypothesisReport],
    meta: dict[str, Any],
) -> Path:
    wb = openpyxl.Workbook()

    # --- Saммари по гипотезам -----------------------------------------------
    ws = wb.active
    ws.title = "Reports"
    ws.append([
        "hypothesis_id", "name", "methodology", "verdict",
        "overall_score", "best_persona", "worst_persona",
        "what_worked", "what_failed", "recommendations",
    ])
    by_id = {h.id: h for h in hypotheses}
    for r in reports:
        h = by_id.get(r.hypothesis_id)
        ws.append([
            r.hypothesis_id,
            h.name if h else "",
            h.methodology if h else "",
            r.verdict,
            r.overall_score,
            r.best_persona,
            r.worst_persona,
            "; ".join(r.what_worked),
            "; ".join(r.what_failed),
            "; ".join(r.recommendations),
        ])

    # --- Гипотезы ------------------------------------------------------------
    ws = wb.create_sheet("Hypotheses")
    ws.append(["id", "name", "methodology", "core_idea", "why_might_work", "key_moves"])
    for h in hypotheses:
        ws.append([h.id, h.name, h.methodology, h.core_idea, h.why_might_work, "; ".join(h.key_moves)])

    # --- Персоны -------------------------------------------------------------
    ws = wb.create_sheet("Personas")
    ws.append(["id", "name", "business_type", "tone", "buying_readiness", "main_objections", "speech_quirks"])
    for p in personas:
        ws.append([
            p.id, p.name, p.business_type, p.tone, p.buying_readiness,
            "; ".join(p.main_objections), p.speech_quirks,
        ])

    # --- Прогоны (вердикты) --------------------------------------------------
    ws = wb.create_sheet("Verdicts")
    headers = [
        "hypothesis_id", "persona_id", "outcome", "avg_score",
        "SPIN", "Sandler", "Challenger", "BANT", "MEDDIC",
        "strengths", "weaknesses",
    ]
    ws.append(headers)
    for v in verdicts:
        score_by = {s.methodology: s.score for s in v.scores}
        ws.append([
            v.hypothesis_id, v.persona_id, v.outcome, v.avg_score,
            score_by.get("SPIN", 0), score_by.get("Sandler", 0),
            score_by.get("Challenger", 0), score_by.get("BANT", 0),
            score_by.get("MEDDIC", 0),
            "; ".join(v.strengths), "; ".join(v.weaknesses),
        ])

    # --- Per-методологический разбор балла (комментарии цензора) ------------
    ws = wb.create_sheet("Scores")
    ws.append(["hypothesis_id", "persona_id", "methodology", "score", "comment"])
    for v in verdicts:
        for s in v.scores:
            ws.append([v.hypothesis_id, v.persona_id, s.methodology, s.score, s.comment])

    # --- Агрегатный разбор по методологиям от Reporter ----------------------
    ws = wb.create_sheet("MethodologySummary")
    ws.append(["hypothesis_id", "methodology", "avg", "note"])
    for r in reports:
        for meth, info in (r.by_methodology or {}).items():
            if not isinstance(info, dict):
                info = {"note": str(info)}
            avg_val = info.get("avg")
            try:
                avg_val = float(avg_val) if avg_val is not None else ""
            except (TypeError, ValueError):
                avg_val = ""
            ws.append([r.hypothesis_id, meth, avg_val, str(info.get("note") or "")])

    # --- Транскрипты симуляций ----------------------------------------------
    ws = wb.create_sheet("Transcripts")
    ws.append(["hypothesis_id", "persona_id", "stop_reason", "n_turns", "transcript"])
    for sim in sims:
        ws.append([
            sim.hypothesis_id, sim.persona_id, sim.stop_reason,
            len(sim.turns), sim.transcript(),
        ])

    # --- Метаданные ----------------------------------------------------------
    ws = wb.create_sheet("Meta")
    ws.append(["key", "value"])
    for k, v in meta.items():
        ws.append([k, str(v)])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
